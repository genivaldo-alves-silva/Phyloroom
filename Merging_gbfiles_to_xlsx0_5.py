##################   Merging files and Genbank file extracting

##when there is no libraries. Prefer to create "environments" to specific libraries (e.g. conda create environment)
##However, you can use as follows; installing libraries out of environments
import os,sys

try:
    import Bio
except ImportError:
    os.system('pip3 install biopython')
	
try:
    import Scipy
except ImportError:
    os.system('pip3 install scipy')
	
try:
    import numpy
except ImportError:
    os.system('pip3 install numpy')
	
try:
    import openpyxl
except ImportError:
    os.system('pip3 install openpyxl')
	
import Bio	
from Bio import SeqIO

import sys
from sys import argv

##################   concatenating gb files###

#All_GBfiles = '{}{}_All_GBfiles.gb'.format(argv[2],argv[3]) #argv[2] = out directory; argv[3] = group name 
group_name = input("what is the group name? ")
All_GBfiles = f"{os.path.dirname(sys.argv[0])}/out/{group_name}_All_GBfiles.gb"


def concatFiles(): # concatenate txt files according to https://stackoverflow.com/questions/13613336/python-concatenate-text-files
    #path = '{}'.format(argv[1]) #argv[1] = input folder
    path = f"{os.path.dirname(sys.argv[0])}/in/"
    files = os.listdir(path)
    for idx, infile in enumerate(files):
        print ("File #" + str(idx) + "  " + infile)
    concat = ''.join([open(path + f).read() for f in files])
    with open(All_GBfiles, "w") as fo:
        fo.write(concat)

if __name__ == "__main__":
    concatFiles()

	
##################   extracting data from Genbank file###

from Bio import SeqIO
import os
input = All_GBfiles
          
#output = '{}{}_Alldata.txt'.format(argv[2],argv[3]) #"F:/Python/room/Alldata.txt"
output = f"{os.path.dirname(sys.argv[0])}/out/{group_name}_Alldata.txt"

handle = open(input)
#possible qualifiers: https://www.insdc.org/submitting-standards/feature-table/#7.4
if not os.path.exists(output): #checks for a pre-existing file with the same name as the output
    for seq_record in SeqIO.parse(handle, "genbank"):
        for seq_feature in seq_record.features:
            if seq_feature.type=="source":
                try: #If you would like your script to run to completion even when that information is not present, you can wrap each access to seq_feature.qualifiers in a try-except block catching KeyError and IndexError
                    strain = seq_feature.qualifiers.get('strain')
                except (KeyError, IndexError):
                    strain = None
                try:
                    db_xref = seq_feature.qualifiers.get('db_xref')
                except (KeyError, IndexError):
                    db_xref = None
                try:
                    type_material = seq_feature.qualifiers.get('type_material')
                except (KeyError, IndexError):
                    type_material = None
                try:
                    specimen_voucher = seq_feature.qualifiers.get('specimen_voucher')
                except (KeyError, IndexError):
                    specimen_voucher = None
                try:
                    isolate = seq_feature.qualifiers.get('isolate')
                except (KeyError, IndexError):
                    isolate = None
                try:
                    environmental_sample = seq_feature.qualifiers.get('environmental_sample')
                except (KeyError, IndexError):
                    environmental_sample = None
                try:
                    culture_collection = seq_feature.qualifiers.get('culture_collection')
                except (KeyError, IndexError):
                    culture_collection = None
                try:
                    country = seq_feature.qualifiers.get('country')
                except (KeyError, IndexError):
                    country = None
                try:
                    isolation_source = seq_feature.qualifiers.get('isolation_source')
                except (KeyError, IndexError):
                    isolation_source = None
                try:
                    host = seq_feature.qualifiers.get('host')
                except (KeyError, IndexError):
                    host = None
                try:
                    order = seq_record.annotations["taxonomy"][-3]
                except (KeyError, IndexError):
                    order = None
                try:
                    family = seq_record.annotations["taxonomy"][-2]
                except (KeyError, IndexError):
                    family = None
                try:
                    genus = seq_record.annotations["taxonomy"][-1]
                except (KeyError, IndexError):
                    genus = None
                try:
                    author = seq_record.annotations['references'][0].authors
                except (KeyError, IndexError):
                    author = None
                try:
                    title = seq_record.annotations['references'][0].title
                except (KeyError, IndexError):
                    title = None
                try:
                    journal = seq_record.annotations['references'][0].journal
                except (KeyError, IndexError):
                    journal = None
                with open(output, "a") as ofile:
                     ofile.write(">{0}xyx{1}xyx{2}xyx{3}xyx{4}xyx{5}xyx{6}xyx{7}xyx{8}xyx{9}xyx{10}xyx{11}xyx{12}xyx{13}xyx{14}xyx{15}xyx{16}xyx{17}xyx{18}xyx{19}xyx{20}zyz\n".format(seq_record.annotations["organism"], order, family, genus, seq_record.name, db_xref, seq_record.description, type_material, strain, specimen_voucher, isolate, environmental_sample, culture_collection, country, isolation_source, host, len(seq_record), author, title, journal, seq_record.seq))
                    
else:
    print ("The output file already seem to exist in the current working directory {0}. Please change the name of the output file".format(os.getcwd())) #error msg

handle.close()

##################   Removing duplicate elements###

#none_duplicate = '{}{}_no_duplicates.txt'.format(argv[2],argv[3])#"F:/Python/room/none_duplicate.txt"
none_duplicate = f"{os.path.dirname(sys.argv[0])}/out/{group_name}_no_duplicates.txt"


import os,sys
 
def Remove(duplicate): # https://www.geeksforgeeks.org/python-remove-duplicates-list/
    final_list = [] 
    for data in duplicate: 
        if data not in final_list: 
            final_list.append(data) 
    return final_list

# Driver Code 
duplicate = open(output) #https://qiita.com/visualskyrim/items/1922429a07ca5f974467

sys.stdout=open(none_duplicate,"w") #https://stackoverflow.com/questions/7152762/how-to-redirect-print-output-to-a-file-using-python/38186276
print(Remove(duplicate))
#sys.stdout.close()
#duplicate.close()

##################   Adding first line for headernames###

#f1 = open('{}{}_no_duplicates.txt'.format(argv[2],argv[3]), 'r')#open("F:/Python/room/none_duplicate.txt", 'r')
f1 = open(f"{os.path.dirname(sys.argv[0])}/out/{group_name}_no_duplicates.txt", "r")
#f2 = open('{}{}_no_duplicates_fw.txt'.format(argv[2],argv[3]), 'w')#open('F:/Python/room/none_duplicate_re.txt', 'w')#output final file
f2 = open(f"{os.path.dirname(sys.argv[0])}/out/{group_name}_no_duplicates_fw.txt", "w")

for j in f1:
   f2.write("\n" + j)


##################   Replace the target string###

#f3 = open('{}{}_no_duplicates_fw.txt'.format(argv[2],argv[3]), 'r')#open('F:/Python/room/none_duplicate_re.txt', 'r')
f3 = open(f"{os.path.dirname(sys.argv[0])}/out/{group_name}_no_duplicates_fw.txt", "r")
#f4 = open('{}{}_no_duplicates_fw2.txt'.format(argv[2],argv[3]), 'w')#open('F:/Python/room/none_duplicate_re2.txt', 'w')
f4 = open(f"{os.path.dirname(sys.argv[0])}/out/{group_name}_no_duplicates_fw2.txt", "w")

for line in f3:
    f4.write(line.replace('xyx','\t').replace('zyz','\n').replace('n", ">','>').replace('[">','>')
	.replace('n"]','').replace("['",'').replace("']",'').replace("\>",'>').replace("n', ",'').replace('n", ','')
	.replace("\'>",'>').replace('\">','>').replace('\>','>').replace("[\'",'').replace('["','').replace("[\'",'')
	.replace("[","").replace("]","").replace("\'",''))


f3.close()
f4.close()


##################   header to xls###


import csv
import openpyxl
from openpyxl import Workbook

csv.field_size_limit()

csv.field_size_limit(100000000)

csv.field_size_limit()

##################   txt to xlsx###

#input_file = '{}{}_no_duplicates_fw2.txt'.format(argv[2],argv[3])#'F:/Python/room/none_duplicate_re2.txt'
input_file = f"{os.path.dirname(sys.argv[0])}/out/{group_name}_no_duplicates_fw2.txt"
#output_file = '{}{}_SpecimensList.xlsx'.format(argv[2],argv[3])#'F:/Python/room/none_duplicate_re2.xlsx'
output_file = f"{os.path.dirname(sys.argv[0])}/out/{group_name}_SpecimensList.xlsx"


headernames = ['Species', 'order', 'family', 'genus', 'GBn', 'db_xref', 'Description', 'type_material', 'strain', 'specimen_voucher', 'isolate', 'environmental_sample', 'culture_collection', 'Country', 'isolation_source', 'host', 'bp', 'authors', 'title', 'journal', 'sequence']

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

with open(input_file, 'r') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)
    for i, j in enumerate(headernames):
        ws.cell(row=1, column = i+1, value= j)

wb.save(output_file)

from openpyxl import load_workbook
#workbook = load_workbook('{}{}_SpecimensList.xlsx'.format(argv[2],argv[3]))
workbook = load_workbook(f"{os.path.dirname(sys.argv[0])}/out/{group_name}_SpecimensList.xlsx")
sheet = workbook.active

sheet.delete_rows(idx=sheet.max_row)#https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

#workbook.save('{}{}_SpecimensList.xlsx'.format(argv[2],argv[3]))
workbook.save(f"{os.path.dirname(sys.argv[0])}/out/{group_name}_SpecimensList.xlsx")

#sys.stdout.close()


import os
import glob

# Path to the directory containing files
directory_path = f"{os.path.dirname(sys.argv[0])}/out/"

# Get a list of all files in the directory
files = glob.glob(directory_path + '*')

# Loop through each file
for f in files:
    # Check if the file does not end with ".xlsx"
    if not f.endswith('.xlsx'):
        # Delete the file
        os.remove(f)


